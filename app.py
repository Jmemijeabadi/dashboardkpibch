from __future__ import annotations

import json
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =========================
# Streamlit Page
# =========================
st.set_page_config(page_title="CRMBI Comercial (Por EJE)", layout="wide")

st.markdown(
    """
    <style>
      .block-container { padding-top: 1.1rem; }
      .kpi-card {
        border-radius: 16px; padding: 14px 16px; background: #ffffff;
        box-shadow: 0 1px 10px rgba(15,23,42,.06);
        border: 1px solid rgba(15,23,42,.06);
      }
      .kpi-title {
        font-size: 12px; letter-spacing: .04em; text-transform: uppercase;
        color: #64748b; font-weight: 800;
      }
      .kpi-value {
        font-size: 34px; font-weight: 900; color: #0f172a;
        line-height: 1.1; margin-top: 6px;
      }
      .kpi-sub { font-size: 12px; color: #64748b; margin-top: 4px; }
      .badge {
        display:inline-block; padding:4px 10px; border-radius:999px;
        font-size:12px; font-weight:700; border:1px solid transparent;
      }
      .badge-green { background:#e8f7ee; color:#0f5132; border-color:#b7e4c7; }
      .badge-yellow{ background:#fff7e6; color:#664d03; border-color:#ffe0a3; }
      .badge-red   { background:#fdecec; color:#842029; border-color:#f5c2c7; }
      .badge-gray  { background:#eef2f7; color:#334155; border-color:#dbe3ef; }
      .muted { color:#64748b; font-size: 13px; }
      .hr { height:1px; background:rgba(15,23,42,.08); margin: 14px 0; }
    </style>
    """,
    unsafe_allow_html=True,
)

# =========================
# CONFIG (KPIs + specialties)
# =========================
KEY_SPECIALTIES = [
    {"id": "ORTOPEDIA",    "match": ["ORTOPEDIA"]},
    {"id": "CARDIOLOGIA",  "match": ["CARDIO"]},
    {"id": "NEUROCIRUGIA", "match": ["NEURO"]},
    {"id": "CX TORAX",     "match": ["TORAX", "TÓRAX"]},
    {"id": "CX GENERAL",   "match": ["CIRUGIA GENERAL", "CIRUGÍA GENERAL", "CX GENERAL"]},
    {"id": "ONCOLOGIA",    "match": ["ONCO"]},
]

KPI_CONFIG: Dict[str, List[Dict[str, Any]]] = {
    "eje1": [
        {
            "id": "eje1_key_specialties_total",
            "indicador": "Atracción médicos nuevos",
            "estrategia": "Actividad por especialidades clave (suma de médicos en: Ortopedia, Cardio, Neuro, Tórax, CX General, Onco).",
            "unit": "médicos",
            "defaultTarget": 25,
            "actualSource": "key_specialties_sum",
        },
        {
            "id": "eje1_eventos_medicos",
            "indicador": "Atracción médicos nuevos",
            "estrategia": "Eventos con contacto 1:1 con médicos (interno/colegios).",
            "unit": "eventos",
            "defaultTarget": 2,
            "actualSource": "not_in_excel",
        },
        {
            "id": "eje1_atenciones_medicos",
            "indicador": "Atención de médicos",
            "estrategia": "Atenciones 1:1 por mes con médicos (seguimiento).",
            "unit": "atenciones",
            "defaultTarget": 80,
            "actualSource": "not_in_excel",
        },
        {
            "id": "eje1_insured_rate",
            "indicador": "Rentabilidad (proxy)",
            "estrategia": "% Seguro / Pacientes (Indicadores).",
            "unit": "%",
            "defaultTarget": 0.35,  # decimal: 0.35 = 35%
            "actualSource": "indicadores_insured_rate",
        },
    ],
    "eje2": [
        {
            "id": "eje2_prospectos_empresas",
            "indicador": "Nuevos convenios",
            "estrategia": "Prospectos de empresas para convenio.",
            "unit": "prospectos",
            "defaultTarget": 6,
            "actualSource": "not_in_excel",
        },
        {
            "id": "eje2_cierres_empresas",
            "indicador": "Nuevos convenios",
            "estrategia": "Cierres de convenio con empresas.",
            "unit": "cierres",
            "defaultTarget": 1,
            "actualSource": "not_in_excel",
        },
        {
            "id": "eje2_eventos_empresas",
            "indicador": "Eventos",
            "estrategia": "Eventos empresariales (ferias / patrocinios / networking).",
            "unit": "eventos",
            "defaultTarget": 1,
            "actualSource": "not_in_excel",
        },
    ],
    "eje3": [
        {
            "id": "eje3_pacientes_seguro",
            "indicador": "Pacientes de seguros",
            "estrategia": "Pacientes con seguro (Indicadores).",
            "unit": "pacientes",
            "defaultTarget": 33,
            "actualSource": "indicadores_insured",
        },
        {
            "id": "eje3_visitas_brokers",
            "indicador": "Pacientes de seguros",
            "estrategia": "Visitas mensuales de agentes y brokers.",
            "unit": "visitas",
            "defaultTarget": 2,
            "actualSource": "not_in_excel",
        },
    ],
}


# =========================
# Session state
# =========================
if "targets" not in st.session_state:
    # { "YYYY-MM": {kpi_id: meta_value} }
    st.session_state.targets = {}

if "status" not in st.session_state:
    st.session_state.status = "Sube el Excel para iniciar."

# =========================
# Utility formatting
# =========================
def badge(text: str, tone: str = "gray"):
    tone_class = {
        "green": "badge badge-green",
        "yellow": "badge badge-yellow",
        "red": "badge badge-red",
        "gray": "badge badge-gray",
    }.get(tone, "badge badge-gray")
    st.markdown(f'<span class="{tone_class}">{text}</span>', unsafe_allow_html=True)


def kpi_card(title: str, value: str, sub: str = ""):
    st.markdown(
        f"""
        <div class="kpi-card">
          <div class="kpi-title">{title}</div>
          <div class="kpi-value">{value}</div>
          <div class="kpi-sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def fmt_int(x: Optional[float]) -> str:
    if x is None:
        return "—"
    try:
        return f"{int(round(float(x))):,}"
    except Exception:
        return "—"


def fmt_pct(x: Optional[float]) -> str:
    if x is None:
        return "—"
    try:
        return f"{float(x) * 100:.1f}%"
    except Exception:
        return "—"


def traffic_light(ratio: Optional[float]) -> Tuple[str, str]:
    if ratio is None:
        return ("⚪ N/D", "gray")
    if ratio >= 1:
        return ("🟢 VERDE", "green")
    if ratio >= 0.8:
        return ("🟡 AMARILLO", "yellow")
    return ("🔴 ROJO", "red")


def download_json_button(label: str, data: Dict[str, Any], filename: str):
    raw = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
    st.download_button(label=label, data=raw, file_name=filename, mime="application/json")


def read_uploaded_json(uploaded_file) -> Dict[str, Any]:
    if uploaded_file is None:
        return {}
    try:
        return json.loads(uploaded_file.read().decode("utf-8"))
    except Exception:
        return {}


# =========================
# Text normalization (accent-insensitive)
# =========================
def norm_text(v: Any) -> str:
    s = "" if v is None else str(v).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()


def parse_month_to_iso(v: Any) -> Optional[str]:
    if v is None:
        return None
    # date/datetime
    if hasattr(v, "year") and hasattr(v, "month"):
        try:
            return f"{int(v.year)}-{int(v.month):02d}-01"
        except Exception:
            pass
    # string
    dt = pd.to_datetime(str(v).strip(), errors="coerce")
    if pd.isna(dt):
        return None
    return f"{dt.year}-{dt.month:02d}-01"


# =========================
# Excel extractors
# =========================
@dataclass
class IndicadoresOut:
    patients: pd.DataFrame
    insured: pd.DataFrame
    months: List[str]
    debug: Dict[str, Any]


def extract_indicadores(ws: Worksheet) -> IndicadoresOut:
    # Read a reasonable region (sheet is small)
    matrix: List[List[Any]] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx > 60:
            break
        matrix.append(list(row))

    # Find "Estadistica"
    header_pos: Optional[Tuple[int, int]] = None
    for r in range(len(matrix)):
        for c in range(min(len(matrix[r]), 80)):
            if norm_text(matrix[r][c]) == "estadistica":
                header_pos = (r, c)
                break
        if header_pos:
            break
    if not header_pos:
        raise ValueError('No encontré "Estadistica" en la hoja Indicadores.')

    hr, hc = header_pos

    # Parse months to the right
    months: List[Tuple[int, str]] = []
    for c in range(hc + 1, len(matrix[hr])):
        iso = parse_month_to_iso(matrix[hr][c])
        if iso:
            months.append((c, iso))
    if not months:
        raise ValueError("No pude parsear meses en Indicadores (fila de 'Estadistica').")

    # Find rows for "No. de Pacientes" and "Seguro"
    p_row = None
    s_row = None
    for r in range(len(matrix)):
        for c in range(min(len(matrix[r]), 30)):
            if norm_text(matrix[r][c]) == norm_text("No. de Pacientes"):
                p_row = r
            if norm_text(matrix[r][c]) == norm_text("Seguro"):
                s_row = r
    if p_row is None:
        raise ValueError('No encontré "No. de Pacientes" en Indicadores.')
    if s_row is None:
        raise ValueError('No encontré "Seguro" en Indicadores.')

    def read_series(row_idx: int) -> pd.DataFrame:
        out = []
        row = matrix[row_idx]
        for c, iso in months:
            if c >= len(row):
                continue
            v = row[c]
            try:
                val = float(v) if v is not None and str(v).strip() != "" else None
            except Exception:
                val = None
            if val is not None:
                out.append({"month": iso, "value": val})
        return pd.DataFrame(out).sort_values("month")

    patients = read_series(p_row)
    insured = read_series(s_row)

    return IndicadoresOut(
        patients=patients,
        insured=insured,
        months=[m for _, m in months],
        debug={"header": {"r": hr + 1, "c": hc + 1}, "p_row": p_row + 1, "s_row": s_row + 1},
    )


@dataclass
class SpecialtiesOut:
    months: List[str]
    series_by_specialty: Dict[str, pd.DataFrame]


def extract_medicos_por_especialidad(ws: Worksheet) -> SpecialtiesOut:
    # Convert to matrix (first ~300 rows enough)
    matrix: List[List[Any]] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx > 300:
            break
        matrix.append(list(row))

    # find row labels cell (Row Labels / Etiquetas de fila)
    header_row = None
    row_labels_col = None
    first_date_col = None
    label_candidates = {"row labels", "rowlabels", "etiquetas de fila", "etiquetasdefila"}

    for r in range(min(80, len(matrix))):
        row = matrix[r]
        for c in range(min(len(row), 40)):
            v = norm_text(row[c])
            if v in label_candidates:
                # find first month column to the right
                for cc in range(c + 1, len(row)):
                    iso = parse_month_to_iso(row[cc])
                    if iso:
                        header_row = r
                        row_labels_col = c
                        first_date_col = cc
                        break
        if header_row is not None:
            break

    if header_row is None or row_labels_col is None or first_date_col is None:
        raise ValueError('No encontré el encabezado del pivote en "Medicos por Especialidad" (Row Labels + fechas).')

    # months
    months: List[Tuple[int, str]] = []
    header_vals = matrix[header_row]
    for c in range(first_date_col, len(header_vals)):
        iso = parse_month_to_iso(header_vals[c])
        if iso:
            months.append((c, iso))
    if not months:
        raise ValueError('No pude parsear meses en "Medicos por Especialidad".')

    # read rows until grand total / total general
    series: Dict[str, pd.DataFrame] = {}
    end_markers = {"grand total", "total general"}

    for r in range(header_row + 1, len(matrix)):
        label = "" if row_labels_col >= len(matrix[r]) else str(matrix[r][row_labels_col] or "").strip()
        if not label:
            continue
        if norm_text(label) in end_markers or any(m in norm_text(label) for m in end_markers):
            break

        vals = []
        row = matrix[r]
        for c, iso in months:
            if c >= len(row):
                val = None
            else:
                v = row[c]
                try:
                    val = float(v) if v is not None and str(v).strip() != "" else None
                except Exception:
                    val = None
            vals.append({"month": iso, "value": val})
        series[label] = pd.DataFrame(vals)

    return SpecialtiesOut(months=[m for _, m in months], series_by_specialty=series)


def series_value_at(df: pd.DataFrame, iso: str) -> Optional[float]:
    if df is None or df.empty:
        return None
    hit = df.loc[df["month"] == iso, "value"]
    if hit.empty:
        return None
    v = hit.iloc[0]
    return float(v) if pd.notna(v) else None


def prev_month_iso(iso: str) -> str:
    d = pd.to_datetime(iso)
    d2 = d - relativedelta(months=1)
    return f"{d2.year}-{d2.month:02d}-01"


def prev_year_iso(iso: str) -> str:
    d = pd.to_datetime(iso)
    d2 = d - relativedelta(years=1)
    return f"{d2.year}-{d2.month:02d}-01"


def find_matching_specialty_label(all_labels: List[str], match_tokens: List[str]) -> Optional[str]:
    tokens = [norm_text(t) for t in match_tokens]
    for lbl in all_labels:
        L = norm_text(lbl)
        if any(t in L for t in tokens):
            return lbl
    return None


def compute_key_specialties_breakdown(spec_series_by_label: Dict[str, pd.DataFrame], iso: str) -> Dict[str, Any]:
    all_labels = list(spec_series_by_label.keys())
    labels, raw_values, values = [], [], []

    for s in KEY_SPECIALTIES:
        lbl = find_matching_specialty_label(all_labels, s["match"])
        labels.append(s["id"])
        if not lbl:
            raw_values.append(None)
            values.append(0.0)
            continue
        v = series_value_at(spec_series_by_label.get(lbl), iso)
        raw_values.append(v if v is not None else None)
        values.append(float(v) if v is not None else 0.0)

    return {"labels": labels, "values": values, "rawValues": raw_values}


@dataclass
class ExcelContext:
    iso: str
    patients: Optional[float]
    insured: Optional[float]
    insured_rate: Optional[float]
    patients_mom: Optional[float]
    patients_yoy: Optional[float]
    insured_mom: Optional[float]
    insured_yoy: Optional[float]
    key_breakdown: Dict[str, Any]


def build_context(ind_patients: pd.DataFrame, ind_insured: pd.DataFrame, spec_series: Dict[str, pd.DataFrame], iso: str) -> ExcelContext:
    p = series_value_at(ind_patients, iso)
    i = series_value_at(ind_insured, iso)
    rate = (i / p) if (i is not None and p is not None and p != 0) else None

    p_prev_m = series_value_at(ind_patients, prev_month_iso(iso))
    p_prev_y = series_value_at(ind_patients, prev_year_iso(iso))
    p_mom = ((p - p_prev_m) / p_prev_m) if (p is not None and p_prev_m not in (None, 0)) else None
    p_yoy = ((p - p_prev_y) / p_prev_y) if (p is not None and p_prev_y not in (None, 0)) else None

    i_prev_m = series_value_at(ind_insured, prev_month_iso(iso))
    i_prev_y = series_value_at(ind_insured, prev_year_iso(iso))
    i_mom = ((i - i_prev_m) / i_prev_m) if (i is not None and i_prev_m not in (None, 0)) else None
    i_yoy = ((i - i_prev_y) / i_prev_y) if (i is not None and i_prev_y not in (None, 0)) else None

    key_breakdown = compute_key_specialties_breakdown(spec_series, iso)

    return ExcelContext(
        iso=iso,
        patients=p,
        insured=i,
        insured_rate=rate,
        patients_mom=p_mom,
        patients_yoy=p_yoy,
        insured_mom=i_mom,
        insured_yoy=i_yoy,
        key_breakdown=key_breakdown,
    )


def get_actual(actual_source: str, ctx: ExcelContext) -> Optional[float]:
    if actual_source == "key_specialties_sum":
        raw = ctx.key_breakdown["rawValues"]
        s = sum(v for v in raw if isinstance(v, (int, float)))
        return s if s > 0 else None
    if actual_source == "indicadores_insured":
        return ctx.insured
    if actual_source == "indicadores_insured_rate":
        return ctx.insured_rate
    return None


def build_kpi_table(eje_key: str, ctx: ExcelContext, targets_for_period: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    for item in KPI_CONFIG[eje_key]:
        target = targets_for_period.get(item["id"], item["defaultTarget"])
        actual = get_actual(item["actualSource"], ctx)

        ratio = (actual / target) if (actual is not None and target not in (None, 0, "")) else None
        sem_text, sem_tone = traffic_light(ratio)
        compliance = fmt_pct(ratio) if ratio is not None else "—"
        actual_disp = fmt_pct(actual) if item["unit"] == "%" else fmt_int(actual)

        # breakdown hint for specialties KPI
        extra = ""
        if item["actualSource"] == "key_specialties_sum":
            parts = []
            for lab, rv in zip(ctx.key_breakdown["labels"], ctx.key_breakdown["rawValues"]):
                parts.append(f"{lab}: {'N/D' if rv is None else int(rv)}")
            extra = " · ".join(parts)

        estrategia = item["estrategia"] + (f" | Desglose: {extra}" if extra else "")

        rows.append({
            "KPI_ID": item["id"],
            "Indicador": item["indicador"],
            "Estrategia": estrategia,
            "Unidad": item["unit"],
            "Meta": float(target) if str(target).strip() != "" else None,
            "Real": actual_disp,
            "Cumplimiento": compliance,
            "Semáforo": sem_text,
            "_tone": sem_tone,
            "_missing": (actual is None and item["actualSource"] == "not_in_excel"),
        })
    return pd.DataFrame(rows)


def apply_targets_edits(df_editor: pd.DataFrame, targets_for_period: Dict[str, Any]) -> Dict[str, Any]:
    for _, r in df_editor.iterrows():
        kpi_id = r.get("KPI_ID")
        meta = r.get("Meta")
        if kpi_id:
            if meta is None or str(meta).strip() == "":
                targets_for_period.pop(kpi_id, None)
            else:
                targets_for_period[kpi_id] = float(meta)
    return targets_for_period


def iso_to_label(iso: str) -> str:
    dt = pd.to_datetime(iso, errors="coerce")
    if pd.isna(dt):
        return iso
    meses = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    return f"{meses[int(dt.month)-1]} {dt.year}"


# =========================
# Cache parsed Excel (fast reruns)
# =========================
@st.cache_data(show_spinner=False)
def parse_excel(file_bytes: bytes) -> Dict[str, Any]:
    # read_only=True to avoid pivot cache overhead / timeouts
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True, keep_links=False)

    # Case-insensitive sheet match
    def get_sheet(wanted: str):
        w = wanted.strip().lower()
        for name in wb.sheetnames:
            if name.strip().lower() == w:
                return wb[name]
        return None

    ws_ind = get_sheet("Indicadores")
    ws_spec = get_sheet("Medicos por Especialidad")

    if ws_ind is None:
        raise ValueError('No existe hoja "Indicadores".')
    if ws_spec is None:
        raise ValueError('No existe hoja "Medicos por Especialidad".')

    ind_out = extract_indicadores(ws_ind)
    spec_out = extract_medicos_por_especialidad(ws_spec)

    return {
        "sheetnames": wb.sheetnames,
        "indicadores": ind_out,
        "specialties": spec_out,
    }


# =========================
# Sidebar (Upload + Period + Targets)
# =========================
st.sidebar.markdown("## CRMBI Comercial")
st.sidebar.caption("Dashboard BI · Excel → KPIs + Tendencias · Metas por periodo")

uploaded = st.sidebar.file_uploader("Excel (.xlsx)", type=["xlsx"])

st.sidebar.divider()
st.sidebar.markdown("### Metas (Targets)")

colA, colB = st.sidebar.columns(2)
with colA:
    reset_btn = st.button("Reset periodo", use_container_width=True)
with colB:
    download_json_button("Export metas", st.session_state.targets, filename="crmbi_targets.json")

import_file = st.sidebar.file_uploader("Import metas (JSON)", type=["json"], key="import_json")
if import_file is not None:
    imported = read_uploaded_json(import_file)
    if isinstance(imported, dict):
        st.session_state.targets = imported
        st.session_state.status = "Metas importadas correctamente (JSON)."
    else:
        st.session_state.status = "No pude importar el JSON (formato inválido)."


# =========================
# Header
# =========================
hl, hr = st.columns([0.78, 0.22], vertical_alignment="center")
with hl:
    st.title("CRMBI Comercial (Por EJE)")
    st.markdown('<div class="muted">Fuente: Indicadores + Médicos por Especialidad · BI Dashboard · Metas editables</div>', unsafe_allow_html=True)
with hr:
    st.markdown("#### Estado")
    st.write(st.session_state.status)

if uploaded is None:
    badge("SIN EXCEL", "gray")
    st.info("Sube el Excel para habilitar KPIs y gráficos.")
    st.stop()

# =========================
# Parse Excel
# =========================
try:
    file_bytes = uploaded.getvalue()
    parsed = parse_excel(file_bytes)
    ind_out: IndicadoresOut = parsed["indicadores"]
    spec_out: SpecialtiesOut = parsed["specialties"]

    available_periods = ind_out.months[:]  # 'YYYY-MM-01'
    if not available_periods:
        raise ValueError("No encontré meses válidos en Indicadores.")

    # default to latest month in Indicadores
    available_periods_sorted = sorted(available_periods)
    default_iso = available_periods_sorted[-1]

    # Period selector (only valid months from your Excel)
    st.sidebar.divider()
    st.sidebar.markdown("### Periodo (desde Excel)")
    iso_selected = st.sidebar.selectbox(
        "Periodo",
        options=available_periods_sorted,
        index=available_periods_sorted.index(default_iso),
        format_func=iso_to_label,
    )

    period_key = iso_selected[:7]  # YYYY-MM

    if reset_btn:
        st.session_state.targets[period_key] = {}
        st.session_state.status = f"Metas reseteadas para {period_key}."

    ctx = build_context(
        ind_patients=ind_out.patients,
        ind_insured=ind_out.insured,
        spec_series=spec_out.series_by_specialty,
        iso=iso_selected,
    )

    badge("EXCEL CARGADO", "green")

except Exception as e:
    badge("ERROR", "red")
    st.error(str(e))
    with st.expander("Debug (error)"):
        st.code(repr(e))
    st.stop()


# =========================
# KPI Cards
# =========================
c1, c2, c3 = st.columns(3)
with c1:
    kpi_card(
        "Pacientes (Indicadores) · periodo",
        fmt_int(ctx.patients),
        f"MoM: {fmt_pct(ctx.patients_mom)} · YoY: {fmt_pct(ctx.patients_yoy)}",
    )
with c2:
    kpi_card(
        "Seguro (Indicadores) · periodo",
        fmt_int(ctx.insured),
        f"MoM: {fmt_pct(ctx.insured_mom)} · YoY: {fmt_pct(ctx.insured_yoy)}",
    )
with c3:
    kpi_card("% Seguro / Pacientes", fmt_pct(ctx.insured_rate), "Seguro ÷ Pacientes")

st.markdown('<div class="hr"></div>', unsafe_allow_html=True)


# =========================
# Charts (Streamlit native)
# =========================
ch1, ch2 = st.columns(2)

with ch1:
    st.subheader("Tendencia: Pacientes")
    dfp = ind_out.patients.copy()
    if dfp.empty:
        st.info("Sin serie de pacientes en el Excel.")
    else:
        dfp2 = dfp.copy()
        dfp2["month"] = pd.to_datetime(dfp2["month"], errors="coerce")
        dfp2 = dfp2.dropna(subset=["month"]).sort_values("month").set_index("month")
        st.line_chart(dfp2["value"])

with ch2:
    st.subheader("Tendencia: Seguro")
    dfi = ind_out.insured.copy()
    if dfi.empty:
        st.info("Sin serie de seguros en el Excel.")
    else:
        dfi2 = dfi.copy()
        dfi2["month"] = pd.to_datetime(dfi2["month"], errors="coerce")
        dfi2 = dfi2.dropna(subset=["month"]).sort_values("month").set_index("month")
        st.line_chart(dfi2["value"])

st.subheader("Especialidades clave (Médicos por Especialidad)")
st.caption("Conteo automático de médicos por especialidad en el mes seleccionado (proxy de actividad).")

df_bar = pd.DataFrame(
    {"Especialidad": ctx.key_breakdown["labels"], "Médicos": ctx.key_breakdown["values"]}
).set_index("Especialidad")
st.bar_chart(df_bar["Médicos"])

st.markdown('<div class="hr"></div>', unsafe_allow_html=True)


# =========================
# Ejes (Tabs)
# =========================
targets_for_period = st.session_state.targets.get(period_key, {})

tabs = st.tabs([
    "EJE 1 · MÉDICOS Y PACIENTES",
    "EJE 2 · EMPRESAS Y COMPASS",
    "EJE 3 · SEGUROS Y BROKERS",
    "DEBUG",
])

eje_meta = {
    "eje1": "Objetivo: Aumentar captación, productividad y rentabilidad de médicos clave.",
    "eje2": "Objetivo: Generar convenios empresariales que alimenten flujo constante de pacientes.",
    "eje3": "Objetivo: Incrementar pacientes asegurados y relaciones con brokers.",
}

def eje_summary(df: pd.DataFrame):
    tones = df["_tone"].value_counts().to_dict()
    greens = tones.get("green", 0)
    yellows = tones.get("yellow", 0)
    reds = tones.get("red", 0)
    nd = tones.get("gray", 0)
    a, b, c, d = st.columns(4)
    with a:
        badge(f"🟢 {greens} Verde", "green")
    with b:
        badge(f"🟡 {yellows} Amarillo", "yellow")
    with c:
        badge(f"🔴 {reds} Rojo", "red")
    with d:
        badge(f"⚪ {nd} N/D", "gray")


for idx, eje_key in enumerate(["eje1", "eje2", "eje3"]):
    with tabs[idx]:
        st.markdown(f"**{eje_meta[eje_key]}**")

        df = build_kpi_table(eje_key, ctx, targets_for_period)
        eje_summary(df)

        missing = df[df["_missing"] == True]
        if len(missing) > 0:
            st.caption("Nota: algunos KPIs están N/D porque no existe ese dato en el Excel (se mantienen como metas manuales).")

        show = df[["KPI_ID", "Indicador", "Estrategia", "Unidad", "Meta", "Real", "Cumplimiento", "Semáforo"]].copy()

        edited = st.data_editor(
            show,
            use_container_width=True,
            hide_index=True,
            disabled=["KPI_ID", "Indicador", "Estrategia", "Unidad", "Real", "Cumplimiento", "Semáforo"],
            column_config={
                "Meta": st.column_config.NumberColumn(
                    "Meta (editable)",
                    help="Para %, usa decimal: 0.35 = 35%.",
                    step=0.01,  # sirve para % y también permite conteos (puedes teclear)
                ),
                "Estrategia": st.column_config.TextColumn("Estrategia", width="large"),
            },
            key=f"editor_{eje_key}_{period_key}",
        )

        b1, b2 = st.columns([0.25, 0.75])
        with b1:
            if st.button("Guardar metas", use_container_width=True, key=f"save_{eje_key}_{period_key}"):
                targets_for_period = apply_targets_edits(edited, targets_for_period)
                st.session_state.targets[period_key] = targets_for_period
                st.session_state.status = f"Metas guardadas para {period_key}."
                st.success("Metas guardadas.")
        with b2:
            st.caption("Tip: Exporta metas (JSON) para conservarlas en Streamlit Cloud o reutilizarlas entre equipos.")

with tabs[3]:
    st.subheader("Debug (Excel)")
    st.caption("Estructura detectada y muestra de especialidades.")
    st.json({
        "sheetnames": parsed.get("sheetnames", []),
        "indicadores": {
            "months_count": len(ind_out.months),
            "months_first_last": [ind_out.months[0], ind_out.months[-1]] if ind_out.months else [],
            "debug": ind_out.debug,
        },
        "medicos_por_especialidad": {
            "months_count": len(spec_out.months),
            "months_first_last": [spec_out.months[0], spec_out.months[-1]] if spec_out.months else [],
            "sample_specialties": list(spec_out.series_by_specialty.keys())[:25],
        },
        "selected_period": period_key,
        "iso_selected": iso_selected,
    })
